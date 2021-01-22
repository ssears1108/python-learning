#!/bin/bash/python3
# this needs to take the different sheets from xlsx file on specific cells 
# and run api calls for VT and compare values of API results and Cylance score. 
# Return and auto populate fields Action and Review notes.  
# https://realpython.com/openpyxl-excel-spreadsheets-python/
# https://support.virustotal.com/hc/en-us/articles/360006819798-API-Scripts-and-client-libraries
import os 
import sys
import time
import requests
import json
import openpyxl
from openpyxl import load_workbook
#from requests.packages.urllib3.exceptions import InsecureRequestWarning
#requests.packages.urllib3.disable_warnings(InsecureRequestWarning)

print("checkpoint 1")
# CONSTANTS   
URL = 'https://www.virustotal.com/vtapi/v2/file/report' #this link not working
# COVERT this to {'apikey': USER_APIKEY, 'resource': None}
PARAMS = {'apikey': 'df5a2bfb15a3e4188db6c26d114a6c11cf1d701b75c18863f310f153efc3bd42', 'resource': None} 
# this is updated in the function below.  
#USER_INPUT = input("Insert Exact File Path")
#USER_APIKEY = input("Your Virus Total API Key")
# is assigning a variable(WORK_DIR) to the os lib object for the directory we are working in, and a variable (PROJECT_DIR) for the folder we are working in.
# REMOVE THIS BEFORE SUBMISSION  
WORK_DIR=os.getcwd()
PROJECT_DIR="soc-automations"
print("checkpoint 2")

# GLOBALS
ITEMS_SEEN = {} # ITEMS_SEEN is an empty dict that allows the values from CY_SCORE, VT_SCORE, and STATUS_SCORE to be stored along with the key hash value
MAL_RESULTS =  {} # this dict will hold file name and 'Quarantine OR Safelist' to be called later for updating the excel document. 
# WORKBOOK = load_workbook(filename = path to file+filename)
print("checkpoint 2.a")
# Create Global USER_INPUT = input("Insert Exact File Path") convert this to WORKBOOK = load.WORKBOOK(filename = USER_INPUT, read_only=True)
WORKBOOK = load_workbook(filename = "{}\\{}\\SPMSSP Rocus Networks - NFR Partner Lab Threats Report (class) 2020-07-24_11-10.xlsx".format(WORK_DIR,PROJECT_DIR)) # removed readOnly = True for test
#WORKBOOK = load_workbook(filename = "{}\\{}\\test.xlsx".format(WORK_DIR,PROJECT_DIR))
print(WORKBOOK.sheetnames) # This displays the available sheets inside the excel document we are working from. 
SHEET_SELECT = input("select sheet to obtain hashes from: ") 
SHEET_SELECTED = WORKBOOK.get_sheet_by_name(SHEET_SELECT)


print(WORKBOOK.get_sheet_by_name(SHEET_SELECT))

print("checkpoint 3")
#iter through rows and columns per sheet for key(SHA256 hash) values (hash, file name, cylance score, status) . 
#row = tuple / use type for more info on what you can do with a tuple consider the fact that unpacking tuples will be required. 
def get_hashes_list(SHEET_SELECTED): 
    for row in SHEET_SELECTED.iter_rows(min_row=2,
        max_row=1000000,
        min_col=1,
        values_only=False): # False = ReadOnlyCell object True = limited values default is False.  
   
        # assuming row is a tuple of ReadOnlyCell objects (see openpyxl docs), return object at index 11 (column = 'SHA2565') and index 3 (column = 'File Name') 
        # and index 15 ('cylance score) and index 5 ('files status')
        SHAHASH = row[11]
        CY_SCORE = row[15]
        FILE_NAME = row[3]
        FILE_STATUS = row[5]
        try:
            assert not isinstance(SHAHASH.value, type(None))," Error: Cell cannot be None"
            yield SHAHASH, CY_SCORE, FILE_NAME, FILE_STATUS #  these are ReadOnlyCell objects  
        except AssertionError as ex: 
            print(ex)
            break 
 
        

# this is the functional code to search for the hash in SHAHASH variable and return the fields from the dict of nested dicts that we care about.  
# hash.row == file name.row will relate these things in logic see docs @ https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/cell/read_only.html
def VT_APICALL(SHEET_SELECTED):
        for SHAHASH, CY_SCORE, FILE_NAME, FILE_STATUS in get_hashes_list(SHEET_SELECTED):
                if SHAHASH.value not in ITEMS_SEEN.keys():
                        PARAMS.update({'resource': SHAHASH.value})
                        response = requests.get(URL,params = PARAMS, verify = False)
                        time.sleep(20)
                        if response.status_code > 200: 
                                print(response.status_code)
                                print("HTTP 204 = Requests Limit Reached. Upgrade to Enterprise Version of Virus Total API.")
                                print(response.status_code)
                        else: 
                            pass
                            VT_RESPONSE = json.loads(response.text)
                        if not isinstance(VT_RESPONSE,list): 
                            VT_SCORE = VT_RESPONSE.get('positives')
                            ITEMS_SEEN.update({SHAHASH.value: dict()}) # add hash as new key, and value as empty to dict to fill in later Added {} to creat new entry. 
                            ITEMS_SEEN[SHAHASH.value]['CY_SCORE'] = CY_SCORE.value
                            ITEMS_SEEN[SHAHASH.value]['VT_SCORE'] = VT_SCORE
                            ITEMS_SEEN[SHAHASH.value]['FILE_NAME'] = FILE_NAME.value
                            ITEMS_SEEN[SHAHASH.value]['FILE_STATUS'] = FILE_STATUS.value 
                            
                            

                        pass
                            
                                

def DETERMINATION_VALUE(SHEET_SELECTED):         
        for HASH, OUTPUT in ITEMS_SEEN.items():
                if HASH not in MAL_RESULTS.keys(): 
                   MAL_RESULTS.update({HASH:"Quarantined"})                
                if OUTPUT["CY_SCORE"] and OUTPUT["VT_SCORE"] >= 0 and OUTPUT["CY_SCORE"]  >= 51 and OUTPUT["VT_SCORE"] >= 3:  
                        MAL_RESULTS[HASH]  ='Quaranitned'
                if OUTPUT["CY_SCORE"]  <= 50 and OUTPUT["VT_SCORE"] <= 3 and OUTPUT["FILE_STATUS"] == 'waived':
                        MAL_RESULTS[HASH] = 'Safelisted'
                else:
                        MAL_RESULTS[HASH] = 'Quarantined'
                 

def WRITE_COLUMNA(SHEET_SELECTED):                                  
        for HASH, OUTPUT in MAL_RESULTS.items(): 
                if MAL_RESULTS.keys() == SHEET_SELECTED.iter_cols(min_col=11,max_col=11,min_row=1,max_row=None):  
                        SHEET_SELECTED.append(A2= OUTPUT) # this needs to point at column name, also needs to be able to iter. 
                        WORKBOOK.save(filename = WORKBOOK) # this closes worksheet
                
        
         


      
                # print(SHAHASH.column, SHAHASH.column_letter, SHAHASH.value)
                # print(CY_SCORE.column, CY_SCORE.column_letter, CY_SCORE.value)
                # print(FILE_NAME.column, FILE_NAME.column_letter, FILE_NAME.value)
                # print(FILE_STATUS.column, FILE_STATUS.column_letter, FILE_STATUS.value)
                # print(VT_SCORE)
                #print(ITEMS_SEEN)
                # proof code works to update dict with nested dicts for values inside. 
                #print(SHEET_SELECTED['A'])

def main():  
       get_hashes_list(SHEET_SELECTED)
       VT_APICALL(SHEET_SELECTED)
       DETERMINATION_VALUE(SHEET_SELECTED)
       WRITE_COLUMNA(SHEET_SELECTED)
       print("Finished")
       
       
       #WRITE_COLUMNA(SHEET_SELECTED)
       
      
if __name__ == "__main__": 
        main()
 
